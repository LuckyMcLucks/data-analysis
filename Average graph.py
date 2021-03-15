import numpy as np
import matplotlib.pyplot as plt
from matplotlib import cm
from matplotlib.colors import ListedColormap, LinearSegmentedColormap
import pandas as pd
import os
from openpyxl import load_workbook

viridis = cm.get_cmap('coolwarm', 256)




def read(x,y):
    wb = load_workbook(filename='Data.xlsx',read_only=True)

    ws = wb['Sheet1']


    data_rows = []
    for row in ws[x:y]:
        data_cols = []
        for cell in row:
            data_cols.append(cell.value)
        data_rows.append(data_cols)
    
    df = pd.DataFrame(data_rows)

    return df.values
    

def plot_examples(x_,y_,name):
    """
    helper function to plot two colormaps
    """
    data = read(x_,y_)
    fig, ax0 = plt.subplots()
    x=[]
    for i in range(166):
        x.append(i/10)
    y=(125,160,200,250,315,400,500,630,800,1000,1250,1600,2000,2500,3150,4000,5000,6300,8000,10000,12500,16000,20000)
    im = ax0.pcolormesh(x,y,data,shading='auto',cmap=viridis)
    ax0.set_title(name)
    fig.colorbar(im, ax=ax0)
    

start = 564
end = 586 
s2 = start+26
s3= s2+26
e2=end+26
e3=e2+26
last = 'FK'
plot_examples('B'+str(start),last+str(end),'22Y')
plot_examples('B'+str(s2),last+str(e2),'23C2')
plot_examples('B'+str(s3),last+str(e3),'24C3')
plt.show()